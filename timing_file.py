import gspread
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import config

filename = config.SERVICE_ACCOUNT
dairy_file = config.DAYS
gbook = config.GBOOK
glist = config.GSHEET
elist = config.ELIST

gc = gspread.service_account(filename)

# Open a sheet from a spreadsheet in one go
spr = gc.open(gbook).worksheet(glist)

# Get all the data from the sheet as a list of lists
data = spr.get_all_values()

# Convert the data to a pandas dataframe
df_google = pd.DataFrame(data[1:], columns=data[0])

# Search for the last not empty element in DAY column
last_index = df_google['DAY'].tolist().index('') - 1
df_google = df_google.loc[:last_index, :]

# Load the Days Diary xlsx workbook
workbook = openpyxl.load_workbook(dairy_file)

# Select the sheet you want to modify
sheet = workbook[elist]

# Save data frame from sheet data
data_days = sheet.values
cols = next(data_days)[0:]
df_days = pd.DataFrame(data_days, columns=cols)

# Rename all columns which have different name because in Google Sheet file modified names are being used
col_names_correct = {}
for i in range(len(cols)):
    if cols[i] != data[0][i]:
        col_names_correct[data[0][i]] = cols[i]
        df_google = df_google.rename(columns=col_names_correct)

# Loop over each column and convert it to the desired data type
for col in df_google.columns:
    # Check if the data type of the column is a string ('object')
    col_dtype = df_days[col].dtype
    # Convert the column to dtype of original dataframe
    if col_dtype == 'datetime64[ns]':
        df_google[col] = pd.to_datetime(df_google[col], format='%d/%m/%Y')
    elif col_dtype == 'float':
        # Apply the transformation to the entire column
        df_google[col] = df_google[col].apply(lambda x: float(x.replace(',', '.')))
        df_google[col] = df_google[col].astype(col_dtype)
    else:
        df_google[col] = df_google[col].astype(col_dtype)


# Create a new worksheet with the same formatting as the existing worksheet
new_worksheet = workbook.create_sheet('Temp')
new_worksheet.sheet_format = sheet.sheet_format
new_worksheet.sheet_properties = sheet.sheet_properties
new_worksheet.page_setup = sheet.page_setup

# Write the updated DataFrame to the worksheet
for r in dataframe_to_rows(df_google, index=False, header=True):
    new_worksheet.append(r)

# Delete all existing rows in the worksheet
sheet.delete_rows(1, sheet.max_row)
for row in new_worksheet.iter_rows():
    sheet.append([cell.value for cell in row])

# Delete the temporary worksheet
workbook.remove(new_worksheet)

# Save the changes to the Excel file
workbook.save(dairy_file)
# Close the workbook
workbook.close()



